from __future__ import annotations

import os
import csv
import json
from typing import List, Dict, Any, Optional, TYPE_CHECKING
from dataclasses import dataclass, field

if TYPE_CHECKING:
    from trace_analyzer import JankAnalysisResult
    from case_config import TestCaseConfig


@dataclass
class CaseSummaryRow:
    case_name: str
    description: str
    duration_s: int
    total_frames: int           # 合计帧（App+SF+其他）
    avg_fps_by_duration: float
    avg_fps_by_frame_interval: float
    max_frame_gap_ms: float
    max_frame_dur_ms: float
    # 整体卡顿等级（合计）- 4级：细微/轻微/明显/严重
    tiny_jank: int
    slight_jank: int
    obvious_jank: int
    severe_jank: int
    jank_ratio_percent: float
    # 以下为带默认值的字段，必须放在所有必填字段之后
    # 帧数按来源拆分
    app_total_frames: int = 0   # App侧帧（用户指定的进程）
    sf_total_frames: int = 0    # SF侧帧（SurfaceFlinger）
    # App侧 & SF侧 平均帧率 (帧数/录制时长，用于区分"成片/预览"两种FPS)
    app_avg_fps: float = 0.0
    sf_avg_fps: float = 0.0
    # App侧（用户进程）卡顿等级 4级
    app_tiny: int = 0
    app_slight: int = 0
    app_obvious: int = 0
    app_severe: int = 0
    # SF侧（SurfaceFlinger）卡顿等级 4级
    sf_tiny: int = 0
    sf_slight: int = 0
    sf_obvious: int = 0
    sf_severe: int = 0
    error: str = ""

    def to_row(self) -> List[Any]:
        return [
            self.case_name,
            self.description,
            self.duration_s,
            self.total_frames,
            self.app_total_frames,
            self.sf_total_frames,
            round(self.avg_fps_by_duration, 2),
            round(self.avg_fps_by_frame_interval, 2),
            round(self.app_avg_fps, 2),
            round(self.sf_avg_fps, 2),
            round(self.max_frame_gap_ms, 2),
            round(self.max_frame_dur_ms, 2),
            self.tiny_jank,
            self.slight_jank,
            self.obvious_jank,
            self.severe_jank,
            self.app_tiny,
            self.app_slight,
            self.app_obvious,
            self.app_severe,
            self.sf_tiny,
            self.sf_slight,
            self.sf_obvious,
            self.sf_severe,
            round(self.jank_ratio_percent, 2),
            self.error,
        ]


SUMMARY_HEADERS = [
    "Case名称", "描述", "录制时长(s)",
    "性能帧数(SF)", "App侧帧数", "SF侧帧数",
    "平均帧率(总帧数/时长)", "平均帧率(按帧间隔)",
    "预览平均帧率(App侧FPS)", "成片平均帧率(SF侧FPS)",
    "最大帧间隔(ms, ts差)", "最大帧时长(ms, dur)",
    # 合计（4级，SF-only）
    "细微卡顿(SF)", "轻微卡顿(SF)", "明显卡顿(SF)", "严重卡顿(SF)",
    # App侧（4级）
    "App侧细微", "App侧轻微", "App侧明显", "App侧严重",
    # SF侧（4级）
    "SF侧细微", "SF侧轻微", "SF侧明显", "SF侧严重",
    "卡顿率(%)", "错误信息"
]


def _zero_bd():
    """返回一个带 tiny/slight/obvious/severe 全0的对象，兼容JankLevelBreakdown不存在的情况"""
    try:
        from trace_analyzer import JankLevelBreakdown
        return JankLevelBreakdown()
    except Exception:
        class _Z:
            tiny = slight = obvious = severe = 0
        return _Z()


def build_summary_row(
    case: "TestCaseConfig",
    result: Optional["JankAnalysisResult"] = None,
    error: str = "",
) -> "CaseSummaryRow":
    if result is not None:
        breakdown = result.jank_level_breakdown
        # 按来源分类
        app_bd = result.per_source_jank_level.get("app") or _zero_bd()
        sf_bd = result.per_source_jank_level.get("sf") or _zero_bd()
        app_frames = result.per_source_total_frames.get("app", 0)
        sf_frames = result.per_source_total_frames.get("sf", 0)
        # App侧 & SF侧 独立FPS（帧数/录制时长），用于区分"预览帧率/成片帧率"
        dur_s = float(case.duration) if case.duration > 0 else 0.0
        app_avg_fps = (app_frames / dur_s) if dur_s > 0 else 0.0
        sf_avg_fps = (sf_frames / dur_s) if dur_s > 0 else 0.0
        return CaseSummaryRow(
            case_name=case.name,
            description=case.description,
            duration_s=case.duration,
            total_frames=result.total_frames,
            app_total_frames=app_frames,
            sf_total_frames=sf_frames,
            avg_fps_by_duration=result.avg_fps_by_duration,
            avg_fps_by_frame_interval=result.avg_fps_by_frame_interval,
            app_avg_fps=app_avg_fps,
            sf_avg_fps=sf_avg_fps,
            max_frame_gap_ms=result.max_frame_gap_ms,
            max_frame_dur_ms=result.max_frame_dur_ms,
            # 整体 4级
            tiny_jank=breakdown.tiny,
            slight_jank=breakdown.slight,
            obvious_jank=breakdown.obvious,
            severe_jank=breakdown.severe,
            jank_ratio_percent=result.jank_ratio * 100,
            # App侧 4级
            app_tiny=getattr(app_bd, "tiny", 0),
            app_slight=getattr(app_bd, "slight", 0),
            app_obvious=getattr(app_bd, "obvious", 0),
            app_severe=getattr(app_bd, "severe", 0),
            # SF侧 4级
            sf_tiny=getattr(sf_bd, "tiny", 0),
            sf_slight=getattr(sf_bd, "slight", 0),
            sf_obvious=getattr(sf_bd, "obvious", 0),
            sf_severe=getattr(sf_bd, "severe", 0),
            error=error,
        )
    return CaseSummaryRow(
        case_name=case.name,
        description=case.description,
        duration_s=case.duration,
        total_frames=0,
        app_total_frames=0,
        sf_total_frames=0,
        avg_fps_by_duration=0.0,
        avg_fps_by_frame_interval=0.0,
        app_avg_fps=0.0,
        sf_avg_fps=0.0,
        max_frame_gap_ms=0.0,
        max_frame_dur_ms=0.0,
        tiny_jank=0,
        slight_jank=0,
        obvious_jank=0,
        severe_jank=0,
        jank_ratio_percent=0.0,
        app_tiny=0,
        app_slight=0,
        app_obvious=0,
        app_severe=0,
        sf_tiny=0,
        sf_slight=0,
        sf_obvious=0,
        sf_severe=0,
        error=error,
    )


def export_summary_csv(rows: List[CaseSummaryRow], output_path: str) -> str:
    """导出汇总表为 CSV（每轮一行 + AVG行）"""
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(SUMMARY_HEADERS)
        for r in rows:
            row_data = r.to_row()
            # AVG行在Case名后加标记
            if r.case_name.endswith("_AVG"):
                row_data[0] = f"★ {r.case_name}"
            writer.writerow(row_data)
    return output_path


def export_summary_excel(rows: List[CaseSummaryRow], output_path: str,
                         detail_reports: Optional[Dict[str, Dict[str, Any]]] = None) -> str:
    """导出汇总表为 Excel (xlsx)。若未安装 openpyxl，则降级为 CSV。
    detail_reports: { case_name: JSON报告dict }，可选，会写入每个case的详情Sheet
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception:
        # 未安装 openpyxl，降级 CSV
        csv_path = os.path.splitext(output_path)[0] + ".csv"
        return export_summary_csv(rows, csv_path)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"

    # 表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(SUMMARY_HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # AVG行样式
    avg_font = Font(bold=True)
    avg_fill = PatternFill("solid", fgColor="FFF2CC")  # 浅黄色背景

    for r in rows:
        ws.append(r.to_row())
        row_idx = ws.max_row
        # AVG行加粗+浅黄背景
        if r.case_name.endswith("_AVG"):
            for cell in ws[row_idx]:
                cell.font = avg_font
                cell.fill = avg_fill

    # 列宽简单自适应（按中文粗略预估）- 26列
    widths = [
        18, 32, 12,     # Case名称, 描述, 录制时长
        10, 10, 10,     # 性能帧数(SF), App侧帧, SF侧帧
        20, 20,         # 两种整体平均帧率
        22, 22,         # 预览帧率(App侧), 成片帧率(SF侧)
        18, 16,         # 最大帧间隔/最大帧时长
        12, 12, 12, 12,  # SF卡顿4级(细微/轻微/明显/严重)
        10, 10, 10, 10,  # App侧4级
        10, 10, 10, 10,  # SF侧4级
        10, 30,         # 卡顿率, 错误信息
    ]
    for i, w in enumerate(widths, start=1):
        col_letter = ""
        n = i
        while n > 0:
            n, rem = divmod(n - 1, 26)
            col_letter = chr(65 + rem) + col_letter
        ws.column_dimensions[col_letter].width = w

    # 详情 Sheet（卡顿等级明细，4级，每轮+AVG）
    if rows:
        detail_ws = wb.create_sheet("卡顿等级明细")
        headers2 = [
            "Case名称",
            "细微(SF)", "轻微(SF)", "明显(SF)", "严重(SF)",
            "App细微", "App轻微", "App明显", "App严重",
            "SF细微", "SF轻微", "SF明显", "SF严重",
        ]
        detail_ws.append(headers2)
        for cell in detail_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for r in rows:
            row_data = [
                r.case_name,
                r.tiny_jank, r.slight_jank, r.obvious_jank, r.severe_jank,
                r.app_tiny, r.app_slight, r.app_obvious, r.app_severe,
                r.sf_tiny, r.sf_slight, r.sf_obvious, r.sf_severe,
            ]
            detail_ws.append(row_data)
            row_idx = detail_ws.max_row
            # AVG行加粗+浅黄背景
            if r.case_name.endswith("_AVG"):
                for cell in detail_ws[row_idx]:
                    cell.font = avg_font
                    cell.fill = avg_fill

    wb.save(output_path)
    return output_path


def export_full_report(rows: List[CaseSummaryRow],
                       output_dir: str,
                       filename_prefix: str = "jank_summary",
                       detail_reports: Optional[Dict[str, Dict[str, Any]]] = None) -> Dict[str, str]:
    """同时导出 JSON、CSV、Excel 三种格式，返回 {格式: 文件路径}"""
    os.makedirs(output_dir, exist_ok=True)
    outputs: Dict[str, str] = {}

    # JSON 汇总
    json_path = os.path.join(output_dir, f"{filename_prefix}.json")
    data = {
        "headers": SUMMARY_HEADERS,
        "rows": [
            dict(zip(SUMMARY_HEADERS, r.to_row()))
            for r in rows
        ],
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    outputs["json"] = json_path

    # CSV
    csv_path = os.path.join(output_dir, f"{filename_prefix}.csv")
    outputs["csv"] = export_summary_csv(rows, csv_path)

    # Excel
    xlsx_path = os.path.join(output_dir, f"{filename_prefix}.xlsx")
    outputs["excel"] = export_summary_excel(rows, xlsx_path, detail_reports)

    return outputs
